from openpyxl import load_workbook


def get_address_list() -> list:
    wb = load_workbook('wallets.xlsx')
    ws = wb['Wallets']
    
    wallet_list = []
    
    for i, wallet in enumerate(ws['A'], start=1):
        if wallet.value and i != 1:
            wallet_list.append(wallet.value)
    
    return wallet_list

def set_balances(worksheet, cell_number: int, balances: dict):
    worksheet[f'B{cell_number}'] = balances['ethereum']
    worksheet[f'C{cell_number}'] = balances['zksync']
    worksheet[f'D{cell_number}'] = balances['arbitrum']
    worksheet[f'E{cell_number}'] = balances['optimism']
    worksheet[f'F{cell_number}'] = balances['scroll']
    
    worksheet[f'G{cell_number}'] = sum(balances.values())

def setup_balances_to_write(all_balances: list):
    wb = load_workbook('wallets.xlsx')
    worksheet = wb['Wallets']
    
    for i, balances in enumerate(all_balances, start=2):
        set_balances(worksheet, i, balances)
    
    wb.save('wallets.xlsx')